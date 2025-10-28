import pandas as pd
import json
import traceback
from audit.models import (
    FileDBMapping,
    PharmacyAuditData,
    DistributorAuditData,
    ProcessLogDetail,
    ErrorSeverity, ErrorLogs, ProcessLogHdr
)
from collections import defaultdict
from audit.models import ErrorLogs, ErrorSeverity
from pharmacy.models import Pharmacy
from threading import Thread
from core.utils import (
    download_file,
    get_default_value_if_null,
    upload_file,
    get_object_or_none,
    get_boto3_client,
)
from .constants import (
    DateFormats,
    FileFormats,
    FileTypes,
    get_output_report_sql,
    get_output_bins_sql,
    get_bin_raw_sql,
)
import traceback
from pharmacy.models import PharmacySoftware
from audit.models import Distributors
import os
from django.conf import settings
from django.apps import apps
from datetime import datetime, timedelta
import pytz
from pharmacy.constants import ProcessingStatusCodes
from pharmacy.utils import get_processing_status
from core.utils import upload_file, get_sql_alchemy_conn
import uuid
import re
from .models import Pharmacy, ProcessLogDetail, FileType
from dateutil import parser
from django.http import JsonResponse, HttpResponse
from wsgiref.util import FileWrapper
import csv
from openpyxl import Workbook, load_workbook
import xlrd
import xlwt
import zipfile
from django.core.files.storage import FileSystemStorage
from django.core.files.base import ContentFile
import io
from botocore.exceptions import ClientError


class ProcessAuditData:
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

    def process_log_details(self, obj):
        try:
            self.file_name = f"{self.process_log.name}.xlsx"
            self.full_file_name = os.path.join(
                os.getcwd(), settings.AUDIT_FILES_LOCATION, self.file_name
            )
            self.writer = pd.ExcelWriter(self.full_file_name, engine="xlsxwriter")
            pd.DataFrame().to_excel(self.writer, sheet_name="output")
            file_location = ""
            print("DEBUG: Entered process_log_details for process:", obj.id)
            details = ProcessLogDetail.objects.filter(process_log=obj.id)
            print("DEBUG: Found", details.count(), "detail records")
            for process_dtl in ProcessLogDetail.objects.filter(
                process_log=obj.id
            ).all():
                file_location = os.path.join(
                    os.getcwd(),
                    settings.AUDIT_FILES_LOCATION,
                    process_dtl.file_name,
                )
                download_file(file_location, process_dtl.file_url)
                if process_dtl.file_type.code == FileTypes.Pharmacy.value:
                    self.process_pharmacy_audit_file(
                        file_location, process_dtl.pharmacy_id
                    )
                elif FileTypes.Distributor.value == process_dtl.file_type.code:
                    self.process_distributor_audit_file(
                        file_location, process_dtl.distributor, self.writer
                    )
        except Exception as e:
            if "invalid input syntax" in str(e).lower():
                # primary_message = e.diag.message_primary
                log_error(
                    process_log=self.process_log,
                    error_message=f"An error occured while generating the report. Please check that all the input data is valid in {os.path.basename(file_location)[8:]}. Error Details: {str(e.orig)}",
                    error_type="Generating report error ",
                    error_severity_code="ER",
                    error_location="trigger_process",
                )
            else:
                log_error(
                    process_log=self.process_log,
                    error_message=f"An error occurred in generating the Report in  {os.path.basename(file_location)}",
                    error_type="Generating report error ",
                    error_severity_code="ER",
                    error_location="trigger_process",
                )

    def trigger_process(self, obj):
        try:
            self.process_log = obj
            self.clean_data_for_process_log()
            if obj:
                self.process_log_details(obj)
                self.generate_output_report()
        except Exception as e:
            obj.status = get_processing_status(ProcessingStatusCodes.Failure.value)
            log_error(
                process_log=self.process_log,
                error_message="An error occurred in generating the Report" + str(e),
                error_type="Generating report error ",
                error_severity_code="ER",
                error_location="trigger_process",
            )
            # obj.log = str(e)
            print(e)
            obj.save()

    def clean_data_for_process_log(self):
        PharmacyAuditData.objects.filter(process_log=self.process_log.id).delete()
        DistributorAuditData.objects.filter(process_log=self.process_log.id).delete()

    def calculate_total_diff(self, x):
        distributor_total = x.sum(axis=1)
        x["Total"] = distributor_total
        pharmacy_quantity = x.index._get_level_values(4)[0]
        x["Difference"] = distributor_total - pharmacy_quantity
        return x

    def highlight(self, val):
        if isinstance(val, (int, float)):
            color = "red" if val < 0 else "green"
            return "color: %s" % color

    def highlight_header(self, val):
        return ["background-color: #808000;" for v in val]

    def format_NDC(self, x):
        if x and x != str(None):
            if len(x) < 11:
                x = "".join(str(0) for i in range(11 - len(x))) + x
            return x[:5] + "-" + x[5:9] + "-" + x[9:11]

    def parse_date(self, x, date_format="%m/%d/%Y"):
        date_formats = DateFormats.formats
        try:
            return datetime.strptime(x, date_format).date()
        except (ValueError, TypeError):
            try:
                return pd.to_datetime(x).date()
            except (ValueError, TypeError):
                for fmt in date_formats:
                    try:
                        return datetime.strptime(x, fmt["value"]).date()
                    except (ValueError, TypeError):
                        continue
        return datetime(1900, 1, 1).date()

    def generate_output_report(self):
        df = pd.read_sql_query(
            get_output_report_sql(
                self.process_log.id,
                self.process_log.pharmacy_from_date,
                self.process_log.pharmacy_to_date,
                self.process_log.distributor_from_date,
                self.process_log.distributor_to_date,
                self.process_log.group,
                self.process_log.pcn,
                self.process_log.bin_number,
            ),
            con=get_sql_alchemy_conn(),
        )
        bin_df = pd.read_sql_query(
            get_output_bins_sql(
                self.process_log.id,
                self.process_log.pharmacy_from_date,
                self.process_log.pharmacy_to_date,
                self.process_log.distributor_from_date,
                self.process_log.distributor_to_date,
                self.process_log.group,
                self.process_log.pcn,
                self.process_log.bin_number,
            ),
            con=get_sql_alchemy_conn(),
        )
        bin_raw_df = pd.read_sql_query(
            get_bin_raw_sql(
                self.process_log.id,
                self.process_log.pharmacy_from_date,
                self.process_log.pharmacy_to_date,
                self.process_log.group,
                self.process_log.pcn,
                self.process_log.bin_number,
            ),
            con=get_sql_alchemy_conn(),
        )
        self.generate_compared_report(df, self.writer, "output")
        self.generate_bin_reports(bin_df, self.writer)
        self.get_bin_raw_data(bin_raw_df, self.writer)
        self.writer.close()
        file_url = f"{settings.AWS_BUCKET}/{self.process_log.name}/{self.file_name}"
        upload_file(self.full_file_name, file_url)
        self.process_log.output_file = file_url
        self.process_log.status = get_processing_status(
            ProcessingStatusCodes.Success.value
        )
        self.process_log.save()
        os.remove(self.full_file_name)

    def generate_bin_reports(self, bin_df, writer):
        bin_groups = bin_df["bgp_name"].unique()
        bin_df["NDC"] = bin_df["NDC"].astype(str).apply(lambda x: self.format_NDC(x))
        bin_df = bin_df.fillna(0)
        for bin_group in sorted(bin_groups):
            df = bin_df[bin_df["bgp_name"] == bin_group]
            df = df.drop(["bgp_name"], axis=1)
            df = df.fillna(0)
            table = pd.pivot_table(
                df,
                values="distributor_quantity",
                index=[
                    "NDC",
                    "Brand",
                    "Drug Name",
                    "Strength",
                    "Dispense Qty in Packs",
                    "Dispense Qty in Units",
                    "Pack",
                    "Total Insurance paid",
                    "Patient Co-pay",
                    "No of RX",
                ],
                columns=["description"],
                aggfunc="sum",
            )

            if 0 in table.columns:
                table = table.drop(0, axis=1)
            table = table.fillna(0)
            table = table.groupby(level=0, group_keys=False).apply(
                self.calculate_total_diff
            )
            table = table.style.map(self.highlight, subset=["Difference"])
            table.to_excel(writer, sheet_name=bin_group)

    def generate_compared_report(self, df, writer, sheet_name):
        df["NDC"] = df["NDC"].astype(str).apply(lambda x: self.format_NDC(x))
        df = df.fillna(0)
        table = pd.pivot_table(
            df,
            values="distributor_quantity",
            index=[
                "NDC",
                "Brand",
                "Drug Name",
                "Strength",
                "Dispense Qty in Packs",
                "Dispense Qty in Units",
                "Pack",
                "Total Insurance paid",
                "Patient Co-pay",
                "No of RX",
            ],
            columns=["description"],
            aggfunc="sum",
        )
        print("DEBUG: Pivot shape:", table.shape)
        print("DEBUG: Pivot index type:", type(table.index))
        print("DEBUG: Pivot columns:", table.columns.tolist())

        if table.empty:
            print("DEBUG: Pivot empty — skipping")
            return
    
        if 0 in table.columns:
            table = table.drop(0, axis=1)
        table = table.fillna(0)
        table = table.groupby(level=0, group_keys=False).apply(
            self.calculate_total_diff
        )
        table = table.style.map(self.highlight, subset=["Difference"])
        table.to_excel(writer, sheet_name=sheet_name)

    def get_bin_raw_data(self, bin_raw_df, writer):
        bin_groups = bin_raw_df["bgp_name"].unique()
        bin_raw_df["NDC"] = (
            bin_raw_df["NDC"].astype(str).apply(lambda x: self.format_NDC(x))
        )
        bin_raw_df = bin_raw_df.fillna(0)
        for bin_group in sorted(bin_groups):
            df = bin_raw_df[bin_raw_df["bgp_name"] == bin_group]
            df = df.drop(["bgp_name"], axis=1)
            df = df.fillna(0)
            df.to_excel(writer, sheet_name=f"{bin_group}_raw_data", index=False)

    def process_pharmacy_audit_file(self, file_location, pharmacy):
        try:
            pharmacy_obj = Pharmacy.objects.get(id=pharmacy)
            self.col_mapping = self.record_to_source_dest_map(
                FileDBMapping.objects.filter(
                    pharmacy_software=pharmacy_obj.software
                ).values("source_col_name", "dest_col_name")
            )
            self.validate_headers(file_location, None, pharmacy)
            pharmacy_file = self.read_file(file_location)
            pharmacy_file = pharmacy_file.rename(columns=self.col_mapping)
            pharmacy_file["pad_ndc"] = pharmacy_file["pad_ndc"].fillna("00000000000")
            pharmacy_file["pad_ndc"] = pharmacy_file["pad_ndc"].apply(
                lambda x: re.sub(
                    "[^a-zA-Z0-9]", "", str(int(x) if isinstance(x, float) else x)
                )
            )
            pharmacy_file["pad_ndc"] = (
                pharmacy_file["pad_ndc"].astype(str).str.zfill(11)
            )
            pharmacy_file["pad_pharmacy"] = pharmacy
            if "pad_date" in pharmacy_file.columns:
                pharmacy_file["pad_date"] = pharmacy_file["pad_date"].apply(
                    lambda x: (
                        x.date()
                        if isinstance(x, (pd.Timestamp, datetime))
                        else self.parse_date(str(x))
                    )
                )
            else:
                pharmacy_file["pad_date"] = datetime(1900, 1, 1).date()
            if "pad_group" in pharmacy_file.columns:
                pharmacy_file["pad_group"] = pharmacy_file["pad_group"].apply(
                    lambda x: str(x).lower() if isinstance(x, str) else x
                )
            if "pad_pcn" in pharmacy_file.columns:
                pharmacy_file["pad_pcn"] = pharmacy_file["pad_pcn"].apply(
                    lambda x: str(x).lower() if isinstance(x, str) else x
                )
            pharmacy_file["pad_quantity"] = pharmacy_file["pad_quantity"].fillna(0)
            pharmacy_file["pad_process_log_id"] = self.process_log.id
            pharmacy_file["pad_ins_bin_number"] = (
                pharmacy_file["pad_ins_bin_number"]
                .apply(lambda x: x.strip() if isinstance(x, str) else x)
                .replace("", 0)
                .fillna(0)
            )
            pharmacy_file["deleted_at"] = datetime.now(tz=pytz.UTC) + timedelta(days=7)
            pharmacy_file["created_at"] = datetime.now(tz=pytz.UTC)
            pharmacy_file["updated_at"] = datetime.now(tz=pytz.UTC)
            pharmacy_file["id"] = pharmacy_file.apply(lambda _: uuid.uuid4(), axis=1)
            pharmacy_file.to_sql(
                "OPT_PAD_PharmacyAuditData",
                get_sql_alchemy_conn(),
                if_exists="append",
                index=False,
                chunksize=10000,
            )
        finally:
            os.remove(file_location)

    def write_file_to_output(self, file_location, writer, distributor_name):
        name, extension = os.path.splitext(file_location)
        data_frame = None
        if extension == FileFormats.CSV.value:
            data_frame = pd.read_csv(file_location)
        elif extension in {FileFormats.XLSX.value, FileFormats.XLS.value}:
            data_frame = pd.read_excel(file_location)
        data_frame.to_excel(writer, sheet_name=distributor_name, index=False)

    def process_distributor_audit_file(self, file_location, distributor, writer):
        try:
            self.col_mapping = self.record_to_source_dest_map(
                FileDBMapping.objects.filter(distributor=distributor).values(
                    "source_col_name", "dest_col_name"
                )
            )
            self.validate_headers(file_location, distributor, None)
            date_type_queryset = FileDBMapping.objects.filter(
                distributor=distributor, dest_col_name="dad_date"
            ).values("date_type")
            self.date_type = (
                date_type_queryset[0]["date_type"]
                if date_type_queryset and date_type_queryset[0]["date_type"] is not None
                else "%m/%d/%Y"
            )
            distributor_file = self.read_file(
                file_location,
            )
            self.write_file_to_output(file_location, writer, distributor.description)
            distributor_file = distributor_file.rename(columns=self.col_mapping)
            distributor_file["dad_ndc"] = distributor_file["dad_ndc"].fillna(
                "00000000000"
            )
            distributor_file["dad_ndc"] = distributor_file["dad_ndc"].apply(
                lambda x: re.sub(
                    "[^a-zA-Z0-9]", "", str(int(x) if isinstance(x, float) else x)
                )
            )
            distributor_file["dad_ndc"] = (
                distributor_file["dad_ndc"].astype(str).str.zfill(11)
            )
            distributor_file["dad_process_log_id"] = self.process_log.id
            distributor_file["dad_distributor"] = distributor.id
            if "dad_date" in distributor_file.columns:
                distributor_file["dad_date"] = distributor_file["dad_date"].apply(
                    lambda x: (
                        x.date()
                        if isinstance(x, (pd.Timestamp, datetime))
                        else self.parse_date(str(x), date_format=self.date_type)
                    )
                )
            else:
                distributor_file["dad_date"] = datetime(1900, 1, 1).date()
            distributor_file["dad_quantity"] = distributor_file["dad_quantity"].replace(
                {",": ""}, regex=True
            )
            distributor_file["dad_quantity"] = (
                distributor_file["dad_quantity"].fillna(0).astype(float)
            )

            distributor_file["deleted_at"] = datetime.now(tz=pytz.UTC) + timedelta(
                days=7
            )
            distributor_file["created_at"] = datetime.now(tz=pytz.UTC)
            distributor_file["updated_at"] = datetime.now(tz=pytz.UTC)
            distributor_file["id"] = distributor_file.apply(
                lambda _: uuid.uuid4(), axis=1
            )

            distributor_file.to_sql(
                "OPT_DAD_DistributorAuditData",
                get_sql_alchemy_conn(),
                if_exists="append",
                index=False,
                chunksize=10000,
            )
        finally:
            os.remove(file_location)

    def record_to_source_dest_map(self, mapping_data):
        return {
            (
                bytes(rec["source_col_name"], "utf-8").decode("unicode_escape")
                if "\\" in rec["source_col_name"]
                else rec["source_col_name"]
            ): rec["dest_col_name"]
            for rec in mapping_data
        }

    def find_header_row(self, df):
        """
        Find the row index where headers are located.
        When multiple source columns map to the same destination (e.g., "NDC", "NDC/UPC", "NDC Number" → "dad_ndc"),
        we need at least ONE of each destination's sources to exist, not ALL sources.
        """
        # Group sources by destination
        dest_to_sources = defaultdict(list)
        for src, dest in self.col_mapping.items():
            dest_to_sources[dest].append(src)

        # Check if at least one source for each destination exists in columns
        def has_all_destinations(cols):
            for dest, sources in dest_to_sources.items():
                # Check if ANY of the source columns for this destination exist
                if not any(flexible_column_match(src, cols) for src in sources):
                    return False
            return True

        # Check if headers are in first row
        if has_all_destinations(list(df.columns)):
            return 0

        # Check subsequent rows
        for i, row in df.iterrows():
            if has_all_destinations(list(row.values)):
                return i + 1

        return 0

    def find_end_row(self, df, header_row):
        end_row_index = None
        ndc_col = self.get_src_size_col("pad_ndc", df)
        for i, row in df[header_row - 1 :].iterrows():
            if all(
                item == "" or pd.isna(item) or item is None or item == 0
                for item in row.values
            ) or (ndc_col and (pd.isna(row[ndc_col]) or row[ndc_col] == "")):
                end_row_index = i - header_row - 1
                break
        return end_row_index

    def read_xl_sheet(self, file_location):
        try:
            xls = pd.ExcelFile(file_location)
            if len(xls.sheet_names) > 0:
                data_frame = pd.read_excel(xls, xls.sheet_names[0])
                header_row = self.find_header_row(data_frame)
                end_row = self.find_end_row(data_frame, header_row)
                if header_row > 0:
                    data_frame = pd.read_excel(
                        xls,
                        xls.sheet_names[0],
                        header=header_row,
                        usecols=self.col_mapping.keys(),
                        nrows=end_row,
                    )
                else:
                    data_frame = pd.read_excel(
                        xls,
                        xls.sheet_names[0],
                        usecols=self.col_mapping.keys(),
                        nrows=end_row,
                    )

                data_frame = data_frame.replace("#MULTIVALUE", None)
                xls.close()
                return data_frame
            return None
        finally:
            if "xls" in locals():
                xls.close()

    def read_csv_file(self, file_location):
        data_frame = pd.read_csv(file_location)
        header_row = self.find_header_row(data_frame)
        end_row = self.find_end_row(data_frame, header_row)
        if header_row > 0:
            data_frame = pd.read_csv(
                file_location,
                header=header_row,
                usecols=self.col_mapping.keys(),
                nrows=end_row,
            )
        else:
            data_frame = pd.read_csv(
                file_location,
                usecols=self.col_mapping.keys(),
                nrows=end_row,
            )

        data_frame = data_frame.replace("#MULTIVALUE", None)
        return data_frame

    def read_file(self, file_location):
        name, extension = os.path.splitext(file_location)
        data_frame = None
        if extension == FileFormats.CSV.value:
            data_frame = self.read_csv_file(file_location)
        elif extension == FileFormats.XLSX.value:
            data_frame = self.read_xl_sheet(file_location)
        elif extension == FileFormats.XLS.value:
            data_frame = self.read_xl_sheet(file_location)
        return data_frame

    def get_col_mappings(self, distributor, pharmacy):
        if distributor:
            self.col_mapping = self.record_to_source_dest_map(
                FileDBMapping.objects.filter(distributor=distributor).values(
                    "source_col_name", "dest_col_name"
                )
            )
        elif pharmacy:
            pharmacy_obj = Pharmacy.objects.get(id=pharmacy)
            self.col_mapping = self.record_to_source_dest_map(
                FileDBMapping.objects.filter(
                    pharmacy_software=pharmacy_obj.software
                ).values("source_col_name", "dest_col_name")
            )

    def get_src_size_col(self, dest_col, df=None):
        """
        Get the source column name that maps to the given destination column.
        If df is provided, returns the source column that actually exists in the dataframe.
        If df is not provided or no match is found, returns the first mapping.
        """
        matching_sources = [src for src, dest in self.col_mapping.items() if dest == dest_col]

        if not matching_sources:
            return None

        # If dataframe provided, find which source column actually exists
        if df is not None:
            df_columns = list(df.columns)
            for src in matching_sources:
                # Use flexible matching to find the column
                matched_col = flexible_column_match(src, df_columns)
                if matched_col:
                    return matched_col

        # Fallback: return first match
        return matching_sources[0]

    def get_file_headers(self, file_location):
        name, extension = os.path.splitext(file_location)
        chunk_size = 100
        max_rows = 1000
        rows_checked = 0
        if extension == FileFormats.CSV.value:
            while rows_checked < max_rows:
                df_chunk = pd.read_csv(
                    file_location, skiprows=rows_checked, nrows=chunk_size
                )
                if not df_chunk.empty:
                    if len(df_chunk.columns) > 2 and not any(
                        pd.isna(val) or "Unnamed" in val for val in df_chunk.columns
                    ):
                        return list(df_chunk.columns)
                    else:
                        for i, row in df_chunk.iterrows():
                            if len(row.values) > 2 and not any(
                                pd.isna(val) for val in row.values
                            ):
                                return list(row.values)
                rows_checked += chunk_size
        elif extension in {FileFormats.XLSX.value, FileFormats.XLS.value}:
            while rows_checked < max_rows:
                df_chunk = pd.read_excel(
                    file_location, skiprows=rows_checked, nrows=chunk_size
                )
                if not df_chunk.empty:
                    if len(df_chunk.columns) > 2 and not any(
                        pd.isna(val) or "Unnamed" in val for val in df_chunk.columns
                    ):
                        return list(df_chunk.columns)
                    else:
                        for i, row in df_chunk.iterrows():
                            if len(row.values) > 2 and not any(
                                pd.isna(val) for val in row.values
                            ):
                                return list(row.values)
                rows_checked += chunk_size
        return set()

    def validate_file(self, file_location, distributor, pharmacy):
        if pharmacy and not distributor:
            data_frame = self.read_file(file_location)
            if (data_frame[self.get_src_size_col("pad_size", data_frame)] == 0).any():
                df_filtered = data_frame[
                    data_frame[self.get_src_size_col("pad_size", data_frame)] == 0
                ]
                invalid_ndc = ",".join(
                    [str(i) for i in df_filtered[self.get_src_size_col("pad_ndc", data_frame)]]
                )
                return False, invalid_ndc

        return True, []

    def validate_headers(self, file_location, distributor, pharmacy):
        self.get_col_mappings(distributor, pharmacy)
        expected_headers = set(self.col_mapping.keys())
        actual_headers = set(self.get_file_headers(file_location))
        if not actual_headers:
            return False, "No headers found in the file."

        required_headers = {value: "" for value in set(self.col_mapping.values())}
        for header in actual_headers:
            if (
                header in expected_headers
                and required_headers[self.col_mapping[header]] == ""
            ):
                required_headers[self.col_mapping[header]] = header
        # missing_headers = {
        #     key: value for key, value in required_headers.items() if value == ""
        # }
        # missing_headers = missing_headers.values()
        # missing_headers = expected_headers - actual_headers
        # if missing_headers:
        #     return False, {", ".join(missing_headers)}
        # self.col_mapping = {value: key for key, value in required_headers.items()}
        # return True, []
        mandatory_columns = {
            "pad_ndc",
            "pad_drug_name",
            "pad_quantity",
            "pad_size",
            "pad_ins_bin_number",
            "dad_ndc",
            "dad_quantity",
            "dad_drug_name",
        }

        missing_headers = {
            key: value
            for key, value in required_headers.items()
            if value == "" and key in mandatory_columns
        }

        for key, value in list(required_headers.items()):
            if value == "" and key not in mandatory_columns:
                required_headers.pop(key)

        missing_headers = list(missing_headers.keys())
        matching_keys = []

        for key, value in self.col_mapping.items():
            if value in missing_headers:
                matching_keys.append(key)

        if missing_headers:
            return False, {", ".join(matching_keys)}

        self.col_mapping = {value: key for key, value in required_headers.items()}

        if hasattr(self, "process_log"):
            unmapped_filters = {
                "pad_date": ["pharmacy_from_date", "pharmacy_to_date"],
                "pad_payment_option": ["payment_method"],
                "pad_claim_status": ["claim_status"],
                "pad_group": ["group"],
                "pad_pcn": ["pcn"],
                "dad_date": ["distributor_from_date", "distributor_to_date"],
            }

            for value in self.col_mapping.values():
                unmapped_filters.pop(value, None)

            for filter_name, attributes in unmapped_filters.items():
                if (pharmacy and filter_name.startswith("pad_")) or (
                    distributor and filter_name == "dad_date"
                ):
                    changed_attributes = []

                    for attr in attributes:
                        if hasattr(self.process_log, attr):
                            current_value = getattr(self.process_log, attr)

                            if attr in ["payment_method", "claim_status"]:
                                if current_value.all():
                                    changed_attributes.append(attr)
                            elif current_value:
                                changed_attributes.append(attr)

                    # Log only if there was a change
                    if changed_attributes:
                        file_name = os.path.basename(file_location)
                        log_error(
                            process_log=self.process_log,
                            error_message=f"The filters {', '.join(changed_attributes)} is missing in {file_name}.",
                            error_type="File Processing Warning",
                            error_severity_code="WA",
                            error_location="validate_headers",
                            error_stack_trace=None,
                        )

        return True, []


def detect_encoding(file_path):
    with open(file_path, "r", encoding="utf-8") as f:
        try:
            first_line = f.readline()
            return "utf-8"
        except:
            return "utf-16"


def detect_delimiter(file_path, encoding="utf-8"):
    with open(file_path, "r", encoding=encoding) as f:
        first_line = f.readline()
        if "," in first_line and "\t" not in first_line:
            return ","
        elif "\t" in first_line and "," not in first_line:
            return "\t"
        else:
            return ","


def get_bad_lines(file_path, delimeter, encoding="utf-8"):
    bad_lines = []
    with open(file_path, "r", encoding=encoding) as f:
        reader = csv.reader(f, delimiter=delimeter)
        for i, line in enumerate(reader):
            if len(line) <= 3:
                bad_lines.append(i)
    return bad_lines


def count_non_blank(row):
    return sum(1 if value is not pd.NA and value != "" else 0 for value in row)


def clean_df(df):
    df = df.dropna(axis=1, how="all")
    changed = False
    df.columns = [
        "" if col is not pd.NA and "Field " in col else col for col in df.columns
    ]
    if count_non_blank(df.columns) > 3:
        return changed, df
    for i, row in enumerate(df.values):
        if count_non_blank(row) > 3:
            changed = True
            df.columns = row
            df = df[(i + 1) :].reset_index(drop=True)
            break
    return changed, df


def clean_file_and_retreive_output_file(full_file_name, output_file_name):
    name, extension = os.path.splitext(full_file_name)
    extension = extension.lower()
    if extension == ".csv":

        encoding = detect_encoding(full_file_name)
        delimiter = detect_delimiter(full_file_name, encoding=encoding)
        bad_lines = get_bad_lines(full_file_name, delimiter, encoding)
        df = pd.read_csv(
            full_file_name,
            skiprows=bad_lines,
            encoding=encoding,
            delimiter=delimiter,
            index_col=False,
        )

    elif extension in [".xls", ".xlsx"]:
        try:
            df = pd.read_excel(
                full_file_name,
                engine="openpyxl" if extension == ".xlsx" else "xlrd",
                dtype=str,
            )
        except Exception as e:
            df = pd.read_excel(full_file_name, engine="openpyxl", dtype=str)
    else:
        raise Exception("Cannot Validate a file with extension:" + extension)
    df = df.dropna(how="all")
    df = df.fillna("")
    df = df.map(lambda x: x.strip() if isinstance(x, str) else x)
    df.columns = ["" if "Unnamed" in col else col for col in df.columns]
    df = df.loc[:, df.notna().any()]
    has_html = True
    while has_html:
        has_html = False
        for column in df.columns:
            if "<div" in column or "<DIV" in column:
                df.columns = df.iloc[0]
                df = df[1:].reset_index(drop=True)
                has_html = True
                break
    df = df.replace(r"^\s*$", pd.NA, regex=True)
    changed = True
    while changed:
        changed, df = clean_df(df)
    for column in df.columns:
        if not pd.isna(column) and "NDC" in column:
            df = df[df[column].notna() & (df[column].astype(str).str.strip() != "")]
    keywords = [
        "Summaries",
        "Sum",
        "thru",
        "Quality Care Products, LLC",
        "Purchases from:",
        "INTERNAL REQUEST",
        "McKESSON CORPORATION",
        "Copyright",
        "Created Date",
        "GRAND TOTAL",
        "TOTAL QTY SHIPPED",
        "Report Total",
        "Powered by",
        "Generated:",
        "records returned",
        "The information provided",
        "Page",
        "orderexpress.cardinalhealth.com",
        "Overall - Summary",
        "Run Date",
        "Report Total:",
        "records returned",
        "Powered by",
    ]

    def is_metadata_row(row):
        row_str = row.astype(str).str.strip()
        contains_keyword = row_str.str.contains(
            "|".join(keywords), case=False, regex=True
        ).any()
        empty_count = row_str.isin(["", "nan", "NaN"]).sum() + row.isna().sum()
        return contains_keyword and empty_count > (len(row) / 2)

    df = df[~df.apply(is_metadata_row, axis=1)]
    df = df.dropna(axis=1, how="all")
    df.columns = [
        f"Blank Named Column {i+1}" if col is pd.NA or not col.strip() else col
        for i, col in enumerate(df.columns)
    ]
    if os.path.exists(output_file_name):
        os.remove(output_file_name)
    if extension == ".csv":
        df.to_csv(output_file_name, index=False)
        return
    if extension == ".xls":
        output_wb = xlwt.Workbook()
        output_wb.add_sheet("Sheet1")
        output_wb.save(output_file_name)
    else:
        output_wb = Workbook()
        output_wb.save(output_file_name)
        output_wb = load_workbook(output_file_name)
        output_sheet = output_wb.create_sheet(title="Sheet1")
    with pd.ExcelWriter(output_file_name, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Sheet1")


def remove_dir_recursive(directory):
    if os.path.exists(directory):
        # Loop through all items in the directory
        for item in os.listdir(directory):
            item_path = os.path.join(directory, item)
            # Check if item is a file or a directory
            if os.path.isfile(item_path) or os.path.islink(item_path):
                os.unlink(item_path)  # Remove the file or symbolic link
            elif os.path.isdir(item_path):
                remove_dir_recursive(item_path)  # Recursively remove the subdirectory
        os.rmdir(directory)  # Remove the now-empty directory


def zip_folder(folder_path, zip_file_name):
    with zipfile.ZipFile(zip_file_name, "w", zipfile.ZIP_DEFLATED) as zipf:
        # Walk through the directory structure
        for root, dirs, files in os.walk(folder_path):
            for file in files:
                # Create the full file path
                file_path = os.path.join(root, file)
                # Add the file to the zip, preserving the folder structure
                arcname = os.path.relpath(file_path, folder_path)
                zipf.write(file_path, arcname)


def batch_process_files(file_location, file_name, file, obj, output_file_name):
    try:
        full_file_name = os.path.join(file_location, file_name)
        name, extension = os.path.splitext(full_file_name)
        if extension == ".zip":
            zip_file_name = os.path.splitext(file.name)[0]
            with zipfile.ZipFile(full_file_name, "r") as zip_ref:
                zip_file, extension = os.path.splitext(name)
                zip_ref.extractall(zip_file)
                os.makedirs(
                    os.path.join(file_location, "cleaned_" + zip_file_name),
                    exist_ok=True,
                )
                try:
                    # threads = []
                    for file_info in zip_ref.infolist():
                        if not file_info.is_dir() and not any(
                            file_info.filename.startswith(excluded)
                            for excluded in settings.ZIP_EXCLUDED_FILES
                        ):
                            cleaned_path = os.path.join(
                                file_location,
                                "cleaned_" + zip_file_name,
                                "cleaned-"
                                + os.path.basename(file_info.filename).lower(),
                            )
 
                            clean_file_and_retreive_output_file(
                                os.path.join(name, file_info.filename),
                                os.path.join(file_location, cleaned_path),
                            )
                    zip_folder(
                        os.path.join(file_location, "cleaned_" + zip_file_name),
                        os.path.join(
                            file_location, "cleaned_" + zip_file_name + ".zip"
                        ),
                    )
                    file_url = f"{settings.AWS_BUCKET}/CleanedFiles/{obj.name}/cleaned_{zip_file_name}.zip"
                    upload_file(
                        os.path.join(
                            file_location, "cleaned_" + zip_file_name + ".zip"
                        ),
                        file_url=file_url,
                    )
                    obj.output_file_url = file_url
                    obj.status = get_processing_status(
                        ProcessingStatusCodes.Success.value
                    )
                    obj.save()
                except Exception as e:
                    raise e
                finally:
                    remove_dir_recursive(zip_file)
                    remove_dir_recursive(
                        os.path.join(file_location, "cleaned_" + zip_file_name)
                    )
        else:
            file_url = (
                f"{settings.AWS_BUCKET}/CleanedFiles/{obj.name}/{output_file_name}"
            )
            output_file_name = os.path.join(file_location, output_file_name)
            clean_file_and_retreive_output_file(full_file_name, output_file_name)
            upload_file(output_file_name, file_url=file_url)
            obj.status = get_processing_status(ProcessingStatusCodes.Success.value)
            obj.output_file_url = file_url
            obj.save()
        os.remove(full_file_name)
        os.remove(os.path.join(file_location, output_file_name))
    except Exception as e:
        obj.status = get_processing_status(ProcessingStatusCodes.Failure.value)
        obj.log = str(e)
        obj.save()
    finally:
        try:
            remove_dir_recursive(file_location)
        except:
            pass


def save_uploaded_file(file, save_path):
    os.makedirs(os.path.dirname(save_path), exist_ok=True)
    with open(save_path, "wb") as f:
        for chunk in file.chunks():
            f.write(chunk)


def unzip_files(zip_file_path, extract_to_folder, process_log):
    try:
        if not zipfile.is_zipfile(zip_file_path):
            error_message = "The provided file is not a valid ZIP file."
            raise ValueError(error_message)
        os.makedirs(extract_to_folder, exist_ok=True)

        with zipfile.ZipFile(zip_file_path, "r") as zip_ref:
            for file in zip_ref.namelist():
                zip_ref.extractall(extract_to_folder)

        extracted_files = []
        for root, dirs, files in os.walk(extract_to_folder):
            for file in files:
                if not any(
                    file.startswith(excluded)
                    for excluded in settings.ZIP_EXCLUDED_FILES
                ):
                    extracted_files.append(file)
        return extracted_files
    except ValueError as e:
        log_error(
            process_log=process_log,
            error_message=error_message,
            error_type="File Validation Error",
            error_severity_code="CR",
            error_location="unzip_files",
        )

    except Exception as e:
        error_message = "An error occured while unzipping the file!"
        log_error(
            process_log=process_log,
            error_message=error_message,
            error_type="File Processing Error",
            error_severity_code="CR",
            error_location="unzip_file",
        )


def validate_required_files(extracted_files, process_log):
    has_pharmacy = False
    has_distributor = False

    if len(extracted_files) < 2:
        error_message = (
            "Validation error: The unzipped folder contains fewer than 2 files."
        )
        log_error(
            process_log=process_log,
            error_message=error_message,
            error_type="Validation Error",
            error_severity_code="ER",
            error_location="unzip_files",
            error_stack_trace=traceback.format_exc(),
        )
    extracted_files = [os.path.splitext(f)[0] for f in extracted_files]
    for file_name in extracted_files:
        file_name = file_name.lower()

        if re.match(r"^pharmacy-.*", file_name):
            pharmacy_name = file_name.split("-")[1]
            if not PharmacySoftware.objects.filter(
                description__iexact=pharmacy_name
            ).exists():
                log_error(
                    process_log=process_log,
                    error_message=f"The pharmacy '{pharmacy_name}' does not exist in the pharmacy software.",
                    error_type="Validation Error",
                    error_severity_code="ER",
                    error_location="validate_required_files",
                )
            else:
                has_pharmacy = True
            continue
        elif re.match(r"^distributor-.*", file_name):
            distributor_name = file_name.split("-")[1]
            if not Distributors.objects.filter(
                description__iexact=distributor_name
            ).exists():
                log_error(
                    process_log=process_log,
                    error_message=f"The Distributor '{distributor_name}' does not exist in the Distributors.",
                    error_type="Validation Error",
                    error_severity_code="ER",
                    error_location="validate_required_files",
                )
            else:
                has_distributor = True
            continue

    missing_files = []
    if not has_pharmacy:
        missing_files.append("pharmacy")
    if not has_distributor:
        missing_files.append("distributor")
    if missing_files:
        error_message = f"Missing required files: {', '.join(missing_files)}."

        log_error(
            process_log=process_log,
            error_message=error_message,
            error_type="Validation Error",
            error_severity_code="ER",
            error_location="validate_required_files",
        )

def handle_zip_file(file, process_log, obj, pharmacy, is_resubmission=False, process_folder=None):
    """
    Handles both initial and reprocess ZIP uploads.
    All files uploaded to S3, local disk is temporary workspace only.
    """
    import shutil
    # Use process_folder if provided, otherwise fall back to process name
    if process_folder is None:
        process_folder = os.path.join(settings.AUDIT_FILES_LOCATION, process_log.name)
        os.makedirs(os.path.join(os.getcwd(), process_folder), exist_ok=True)

    temp_dir = os.path.join(os.getcwd(), "temp_files", f"process_{process_log.id}")
    extract_to_folder = os.path.join(temp_dir, "extracted")
    remove_dir_recursive(temp_dir)
    os.makedirs(extract_to_folder, exist_ok=True)

    created_files = []
    failed_files = []
    pharmacy_processed = 0
    pharmacy_failed = 0
    distributor_processed = 0
    distributor_failed = 0

    try:
        if is_resubmission:
            print(f"♻️ Starting reprocess for Process ID: {process_log.id}")

            process_log.status = get_processing_status(ProcessingStatusCodes.Inprogress.value)
            process_log.save(update_fields=["status"])

            ErrorLogs.objects.filter(process_log=process_log).delete()

            failed_list = []
            if process_log.failed_files_json and process_log.failed_files_json != "[]":
                try:
                    failed_list = json.loads(process_log.failed_files_json)
                except Exception:
                    failed_list = []

            #  Delete old failed files from S3 during reprocess
            s3_client = get_boto3_client()
            audit_dir = os.path.join(os.getcwd(), process_folder)
            
            for failed_name in failed_list:
                possible_names = [failed_name, f"cleaned_{failed_name}"]
                
                # Delete DB entries
                ProcessLogDetail.objects.filter(
                    process_log=process_log, file_name__in=possible_names
                ).delete()

                # Delete from S3 FIRST
                for name in possible_names:
                    s3_key = f"{process_log.name}/{name}"
                    try:
                        s3_client.delete_object(Bucket=settings.AWS_BUCKET, Key=s3_key)
                        print(f"🗑️ Deleted from S3: {s3_key}")
                    except Exception as e:
                        print(f" Could not delete from S3 {s3_key}: {e}")

                # Delete local files
                for name in possible_names:
                    fpath = os.path.join(audit_dir, name)
                    if os.path.exists(fpath):
                        try:
                            os.remove(fpath)
                            print(f"🧹 Removed previously failed file: {name}")
                        except Exception as e:
                            print(f" Could not remove failed file {name}: {e}")

            pharmacy_processed = process_log.pharmacy_processed_count or 0
            distributor_processed = process_log.distributor_processed_count or 0
        else:
            process_log.failed_files_json = "[]"
            process_log.failed_count = 0
            process_log.pharmacy_processed_count = 0
            process_log.pharmacy_failed_count = 0
            process_log.distributor_processed_count = 0
            process_log.distributor_failed_count = 0
            process_log.save()

        # Save uploaded ZIP
        temp_zip_path = os.path.join(temp_dir, "temp.zip")
        if isinstance(file, io.BytesIO):
            with open(temp_zip_path, "wb") as temp_zip_file:
                temp_zip_file.write(file.getvalue())
        else:
            fs = FileSystemStorage(temp_dir)
            with file.open("rb") as f:
                content = f.read()
                fs.save(file.name, ContentFile(content))
            temp_zip_path = os.path.join(temp_dir, file.name)

        # Extract files
        extracted_files = unzip_files(temp_zip_path, extract_to_folder, process_log)

        if not is_resubmission:
            validate_required_files(extracted_files, process_log)

        entries = os.listdir(extract_to_folder)

        # Clean & register files
        for extracted_file in extracted_files:
            if len(entries) == 1 and os.path.isdir(os.path.join(extract_to_folder, entries[0])):
                folder_name = entries[0]
                base_name = os.path.join(extract_to_folder, folder_name, extracted_file)
            else:
                base_name = os.path.join(extract_to_folder, extracted_file)

            original_name = os.path.basename(base_name)
            output_file = os.path.join(os.getcwd(), process_folder, f"cleaned_{original_name}")   
            created_files.append(output_file)

            is_pharmacy_file = "pharmacy" in original_name.lower()
            is_distributor_file = "distributor" in original_name.lower()

            try:
                clean_file_and_retreive_output_file(base_name, output_file)
            except Exception as e:
                failed_files.append(original_name)
                if is_pharmacy_file:
                    pharmacy_failed += 1
                elif is_distributor_file:
                    distributor_failed += 1
                    
                log_error(
                    process_log=process_log,
                    error_message=f"Error cleaning {original_name}: {e}",
                    error_type="File Cleaning Error",
                    error_severity_code="ER",
                    error_location="handle_zip_file",
                    error_stack_trace=traceback.format_exc(),
                )
                continue

            file_name = os.path.basename(output_file)
            name, ext = os.path.splitext(file_name)
            
            if ext.lower() in {".xlsx", ".xls", ".csv"}:
                columns = extract_column_names(output_file)
                
                # S3-FIRST: Use full S3 path for file_url
                full_file_url = f"{settings.AWS_BUCKET}/{process_log.name}/{file_name}"

                # DISTRIBUTOR FILE
                if is_distributor_file:
                    distributor_name = name.replace("cleaned_", "").split("-")[1]
                    if Distributors.objects.filter(description__icontains=distributor_name).exists():
                        distributor = Distributors.objects.filter(
                            description__icontains=distributor_name
                        ).first().id
                        
                        if check_column_mappings(extracted_file, distributor, columns, process_log):
                            ProcessLogDetail.objects.filter(
                                process_log=process_log, file_name=file_name
                            ).delete()
                            ProcessLogDetail.objects.create(
                                file_type=FileType.objects.get(description="Distributor"),
                                file_name=file_name,
                                process_log=process_log,
                                file_url=full_file_url,  # Full S3 path
                                distributor=get_object_or_none(
                                    Distributors,
                                    pk=get_default_value_if_null(distributor, None),
                                ),
                            )
                            distributor_processed += 1
                        else:
                            failed_files.append(original_name)
                            distributor_failed += 1
                    else:
                        failed_files.append(original_name)
                        distributor_failed += 1
                        log_error(
                            process_log=process_log,
                            error_message=f"Distributor '{distributor_name}' not found.",
                            error_type="Validation Error",
                            error_severity_code="ER",
                            error_location="handle_zip_file",
                        )

                # PHARMACY FILE
                elif is_pharmacy_file:
                    pharmacy_name = name.replace("cleaned_", "").split("-")[1]
                    if PharmacySoftware.objects.filter(description__icontains=pharmacy_name).exists():
                        software = PharmacySoftware.objects.filter(
                            description__icontains=pharmacy_name
                        ).first().id
                        
                        if check_Phamacy_column_mappings(extracted_file, software, columns, process_log):
                            ProcessLogDetail.objects.filter(
                                process_log=process_log, file_name=file_name
                            ).delete()
                            ProcessLogDetail.objects.create(
                                file_type=FileType.objects.get(description="Pharmacy"),
                                file_name=file_name,
                                process_log=process_log,
                                file_url=full_file_url,  # Full S3 path
                                pharmacy=get_object_or_none(
                                    Pharmacy, pk=get_default_value_if_null(pharmacy, None)
                                ),
                            )
                            process_audit_data = ProcessAuditData()
                            process_audit_data.validate_headers(output_file, False, pharmacy)
                            is_valid_file, invalid_ndcs = process_audit_data.validate_file(
                                output_file, False, pharmacy
                            )
                            if not is_valid_file:
                                failed_files.append(original_name)
                                pharmacy_failed += 1
                                log_error(
                                    process_log=process_log,
                                    error_message=f"Invalid NDCs {invalid_ndcs} in {original_name}",
                                    error_type="Validation Error",
                                    error_severity_code="ER",
                                    error_location="validate_file",
                                )
                            else:
                                pharmacy_processed += 1
                        else:
                            failed_files.append(original_name)
                            pharmacy_failed += 1

        # Upload ALL files to S3 FIRST (successful + failed for forensics)
        upload_process_folder_to_s3(process_log.name)
        
        # THEN clean failed files from local audit folder (after S3 upload)
        audit_dir = os.path.join(os.getcwd(), process_folder)
        for failed in failed_files:
            possible_names = [failed, f"cleaned_{failed}"]
            for name in possible_names:
                fpath = os.path.join(audit_dir, name)
                if os.path.exists(fpath):
                    try:
                        os.remove(fpath)
                        print(f"🧹 Removed failed file from local audit folder: {fpath}")
                    except Exception as e:
                        print(f" Could not remove failed file {fpath}: {e}")

        # Trigger final report
        process_audit_data = ProcessAuditData()
        error_severity = ErrorSeverity.objects.get(code="WA").id
        
        # Don't set status here - trigger_process handles it
        if not ErrorLogs.objects.filter(process_log=process_log).exclude(
            error_severity=error_severity
        ).exists():
            # trigger_process sets status to Success in generate_output_report()
            process_audit_data.trigger_process(obj)
        else:
            # Only set failure status if there are errors
            process_log.status = get_processing_status(ProcessingStatusCodes.Failure.value)

        # Update counts
        process_log.failed_files_json = json.dumps(failed_files)
        process_log.failed_count = len(failed_files)
        process_log.pharmacy_processed_count = pharmacy_processed
        process_log.pharmacy_failed_count = pharmacy_failed
        process_log.distributor_processed_count = distributor_processed
        process_log.distributor_failed_count = distributor_failed
        
        # Save (status already handled by trigger_process if successful)
        update_fields = [
            "pharmacy_processed_count", 
            "pharmacy_failed_count",
            "distributor_processed_count",
            "distributor_failed_count",
            "failed_count", 
            "failed_files_json"
        ]
        
        # Only include status if there were errors
        if ErrorLogs.objects.filter(process_log=process_log).exclude(
            error_severity=error_severity
        ).exists():
            update_fields.append("status")
        
        process_log.save(update_fields=update_fields)

        # Clean up local files after S3 upload
        cleanup_local_process_folder(process_log.name)

    except Exception as e:
        process_log.status = get_processing_status(ProcessingStatusCodes.Failure.value)
        process_log.failed_files_json = json.dumps(failed_files)
        process_log.failed_count = len(failed_files)
        process_log.pharmacy_processed_count = pharmacy_processed
        process_log.pharmacy_failed_count = pharmacy_failed
        process_log.distributor_processed_count = distributor_processed
        process_log.distributor_failed_count = distributor_failed
        
        process_log.save(
            update_fields=[
                "status", 
                "pharmacy_processed_count",
                "pharmacy_failed_count",
                "distributor_processed_count",
                "distributor_failed_count",
                "failed_files_json", 
                "failed_count"
            ]
        )
        log_error(
            process_log=process_log,
            error_message=str(e),
            error_type="Zip Processing Error",
            error_severity_code="ER",
            error_location="handle_zip_file",
            error_stack_trace=traceback.format_exc(),
        )

    finally:
        remove_dir_recursive(temp_dir)
        if os.path.exists(temp_zip_path):
            os.remove(temp_zip_path)
    
def cleanup_temp_dir(path):
    """Clean extracted temp directory to avoid collisions."""
    if os.path.exists(path):
        for root, dirs, files in os.walk(path, topdown=False):
            for f in files:
                try:
                    os.remove(os.path.join(root, f))
                except Exception:
                    pass
            for d in dirs:
                try:
                    os.rmdir(os.path.join(root, d))
                except Exception:
                    pass

def log_error(
    error_message,
    process_log=None,
    error_type=None,
    error_severity_code=None,
    error_location=None,
    user_context=None,
    error_stack_trace=None,
    created_by=None,
):
    try:
        error_severity = None
        if error_severity_code:
            error_severity = ErrorSeverity.objects.filter(
                code=error_severity_code
            ).first()
        # if process_log:
        #     error_message = f"[ProcessLog ID: {process_log.id}] {error_message}"
        if error_stack_trace is None:
            error_stack_trace = traceback.format_exc()
        
        # --- Detect failed filename in message ---
        failed_filename = None
        if " in " in error_message and "." in error_message:
            parts = error_message.split(" in ")
            if len(parts) > 1:
                candidate = parts[-1].strip().split()[0]
                if any(candidate.lower().endswith(ext) for ext in [".xlsx", ".xls", ".csv"]):
                    failed_filename = candidate

        ErrorLogs.objects.create(
            error_message=error_message,
            error_type=error_type,
            error_severity=error_severity,
            error_location=error_location,
            user_context=user_context,
            process_log=process_log,
            error_stack_trace=error_stack_trace,
            created_by=created_by,
        )
        
        # --- Also store failed filenames in ProcessLogHdr ---
        if process_log:
            try:
                plog = ProcessLogHdr.objects.get(pk=process_log.id)

                # Update failed files list if we detected a filename
                if failed_filename:
                    existing = []
                    if plog.failed_files_json:
                        try:
                            existing = json.loads(plog.failed_files_json)
                            if not isinstance(existing, list):
                                existing = []
                        except Exception:
                            existing = []

                    if failed_filename not in existing:
                        existing.append(failed_filename)
                        plog.failed_files_json = json.dumps(existing)

                # Determine if this is a pharmacy or distributor error
                # Check multiple indicators: filename, error message, and error location
                is_pharmacy_error = False
                is_distributor_error = False

                # 1. Check filename if available
                if failed_filename:
                    filename_lower = failed_filename.lower()
                    # Common pharmacy file patterns
                    if any(pattern in filename_lower for pattern in ['pharmacy', 'rx', 'pharm', 'dispensing', 'prescription']):
                        is_pharmacy_error = True
                    # Common distributor file patterns
                    elif any(pattern in filename_lower for pattern in ['distributor', 'dist', 'wholesaler', 'purchase', 'invoice']):
                        is_distributor_error = True

                # 2. Check error message for file context if filename check didn't work
                if not is_pharmacy_error and not is_distributor_error:
                    error_msg_lower = error_message.lower()
                    if any(pattern in error_msg_lower for pattern in ['pharmacy', 'rx', 'dispensing', 'prescription']):
                        is_pharmacy_error = True
                    elif any(pattern in error_msg_lower for pattern in ['distributor', 'dist', 'wholesaler', 'purchase', 'invoice']):
                        is_distributor_error = True

                # 3. Fall back to error_location as last resort
                if not is_pharmacy_error and not is_distributor_error and error_location:
                    error_loc_lower = error_location.lower()
                    if 'pharmacy' in error_loc_lower or 'rx' in error_loc_lower:
                        is_pharmacy_error = True
                    elif 'distributor' in error_loc_lower:
                        is_distributor_error = True

                update_fields = []

                # Increment the appropriate failed count
                if is_pharmacy_error:
                    plog.pharmacy_failed_count = (plog.pharmacy_failed_count or 0) + 1
                    update_fields.append("pharmacy_failed_count")
                elif is_distributor_error:
                    plog.distributor_failed_count = (plog.distributor_failed_count or 0) + 1
                    update_fields.append("distributor_failed_count")
                else:
                    # If we can't determine, increment legacy failed_count
                    plog.failed_count = (plog.failed_count or 0) + 1
                    update_fields.append("failed_count")

                if failed_filename and "failed_files_json" not in update_fields:
                    update_fields.append("failed_files_json")

                if update_fields:
                    plog.save(update_fields=update_fields)

            except Exception as inner_ex:
                print(f"Warning: could not update failed counts: {inner_ex}")

    except Exception as e:

        print(f"Failed to log error: {e}")

def extract_column_names(file_path, sheet_name=0):
    file_extension = os.path.splitext(file_path)[1].lower()
    if file_extension == ".csv":
        df = pd.read_csv(file_path)
    elif file_extension in {".xls", ".xlsx"}:
        df = pd.read_excel(file_path, sheet_name=sheet_name)
    # df = pd.read_excel(file_path, sheet_name=sheet_name)
    column_names = list(df.columns)

    return column_names

def flexible_column_match(expected_col_name, actual_columns):
    """
    Match a column name against a list of actual column names with case-insensitive comparison.

    The FileDBMapping table should contain all valid variations (e.g., "NDC", "NDC/UPC", "NDC Number"),
    so we only need case-insensitive exact matching here.

    Returns the matching column name from actual_columns, or None if no match found.
    """
    if not expected_col_name or not actual_columns:
        return None

    # Normalize the expected column name (strip whitespace, lowercase)
    expected_normalized = expected_col_name.strip().lower()

    # Case-insensitive exact match
    for col in actual_columns:
        if col.strip().lower() == expected_normalized:
            return col

    return None


def check_column_mappings(extrated_file, distributor, columns, process_log):
    mappings = FileDBMapping.objects.filter(distributor=distributor).values(
        "source_col_name", "dest_col_name"
    )
    source_to_dest_map = defaultdict(list)
    for mapping in mappings:
        dest_col_name = mapping["dest_col_name"]
        source_col_name = mapping["source_col_name"]
        source_to_dest_map[dest_col_name].append(source_col_name)

    grouped_destinations = {
        dest_col_name: False for dest_col_name in source_to_dest_map
    }
    for mapping in mappings:
        source_col_name = mapping["source_col_name"]
        dest_col_name = mapping["dest_col_name"]
        # Use flexible column matching instead of exact match
        matched_col = flexible_column_match(source_col_name, columns)
        if matched_col:
            grouped_destinations[dest_col_name] = True
    if not grouped_destinations.get("dad_date", False):
        log_error(
            process_log=process_log,
            error_message=(
                f"The column '{', '.join(source_to_dest_map['dad_date'])}' is missing in the {extrated_file}."
                if source_to_dest_map["dad_date"]
                else f"The date column is missing in the {extrated_file}."
            ),
            error_type="File Processing Warning",
            error_severity_code="WA",
            error_location="Check_column_mappings",
            error_stack_trace=None,
        )
        grouped_destinations.pop("dad_date", None)
    missing_dest_columns = [
        dest for dest, valid in grouped_destinations.items() if not valid
    ]
    for dest in missing_dest_columns:
        missing_sources_for_dest = source_to_dest_map[dest]
        if missing_sources_for_dest:
            error_message = (
                f"At least one of the following columns is required in {extrated_file} : {', '.join(missing_sources_for_dest)} "
                if len(missing_sources_for_dest) > 1
                else f"The source column '{missing_sources_for_dest[0]}' is required in {extrated_file}."
            )
            log_error(
                process_log=process_log,
                error_message=error_message,
                error_type="File Processing Error",
                error_severity_code="ER",  # ER for Error
                error_location="Check_column_mappings",
                error_stack_trace=None,
            )
    all_valid = not missing_dest_columns
    # all_valid = True

    return all_valid

def check_Phamacy_column_mappings(
    extracted_file, pharmacysoftware, columns, process_log
):

    required_columns = {
        "pad_ndc",
        "pad_drug_name",
        "pad_quantity",
        "pad_size",
        "pad_ins_bin_number",
    }
    mappings = FileDBMapping.objects.filter(pharmacy_software=pharmacysoftware).values(
        "source_col_name", "dest_col_name"
    )
    # dest_to_source_map = {mapping["dest_col_name"]: mapping["source_col_name"] for mapping in mappings}
    grouped_destinations = {mapping["dest_col_name"]: False for mapping in mappings}
    dest_to_source_map = defaultdict(list)
    for mapping in mappings:
        source_col_name = mapping["source_col_name"]
        dest_col_name = mapping["dest_col_name"]
        dest_to_source_map[dest_col_name].append(source_col_name)

        # Use flexible column matching instead of exact match
        matched_col = flexible_column_match(source_col_name, columns)
        if matched_col:
            grouped_destinations[dest_col_name] = True
    missing_required_columns = [
        dest for dest in required_columns if not grouped_destinations.get(dest, False)
    ]
    missing_optional_columns = [
        dest
        for dest, valid in grouped_destinations.items()
        if not valid and dest not in required_columns
    ]

    for dest in missing_required_columns:
        missing_sources_for_dest = dest_to_source_map[dest]
        if missing_sources_for_dest:
            error_message = (
                f"At least one of the following required columns is missing in {extracted_file}: {', '.join(missing_sources_for_dest)}."
                if len(missing_sources_for_dest) > 1
                else f"The required column '{missing_sources_for_dest[0]}' is missing in {extracted_file}."
            )
            log_error(
                process_log=process_log,
                error_message=error_message,
                error_type="File Processing Error",
                error_severity_code="ER",
                error_location="Check_column_mappings",
                error_stack_trace=None,
            )

    for dest in missing_optional_columns:
        missing_sources_for_dest = dest_to_source_map[dest]
        if missing_sources_for_dest:
            error_message = (
                f"Validation warning: At least one of the following optional columns is missing in {extracted_file}: {', '.join(missing_sources_for_dest)}."
                if len(missing_sources_for_dest) > 1
                else f"Validation warning: The optional column '{missing_sources_for_dest[0]}' is missing in {extracted_file}."
            )
            log_error(
                process_log=process_log,
                error_message=error_message,
                error_type="File Processing Warning",
                error_severity_code="WA",
                error_location="Check_column_mappings",
                error_stack_trace=None,
            )

    all_valid = not missing_required_columns
    return all_valid

def upload_process_folder_to_s3(process_name):
    """
    Upload all files in process folder to S3, skip duplicates.
    Maintains same folder structure as local.
    """
    s3_client = get_boto3_client()
    local_folder = os.path.join(os.getcwd(), settings.AUDIT_FILES_LOCATION, process_name)
    if not os.path.exists(local_folder):
        print(f"⚠️ Local folder does not exist: {local_folder}")
        return

    uploaded_count = 0
    skipped_count = 0

    for root, dirs, files in os.walk(local_folder):
        for file_name in files:
            local_file_path = os.path.join(root, file_name)
            s3_key = f"{process_name}/{file_name}"
            
            # Check if file already exists in S3
            try:
                s3_client.head_object(Bucket=settings.AWS_BUCKET, Key=s3_key)
                print(f"⏭️ Skipping {file_name} (already in S3)")
                skipped_count += 1
                continue
            except ClientError:
                pass
            
            # Upload file to S3
            try:
                s3_client.upload_file(local_file_path, settings.AWS_BUCKET, s3_key)
                print(f"✅ Uploaded {file_name} to S3")
                uploaded_count += 1
            except Exception as e:
                print(f"❌ Failed to upload {file_name}: {e}")

    print(f" Upload complete: {uploaded_count} uploaded, {skipped_count} skipped")
    
def cleanup_s3_folder(process_name):
    """
    Delete entire S3 folder when process deleted from UI.
    """
    s3_client = get_boto3_client()
    prefix = f"{process_name}/"
    try:
        response = s3_client.list_objects_v2(
            Bucket=settings.AWS_BUCKET,
            Prefix=prefix
        )
        
        if 'Contents' not in response:
            print(f"No files found in S3 for process: {process_name}")
            return
        
        objects_to_delete = [{'Key': obj['Key']} for obj in response['Contents']]
        
        if objects_to_delete:
            s3_client.delete_objects(
                Bucket=settings.AWS_BUCKET,
                Delete={'Objects': objects_to_delete}
            )
            print(f"🗑️ Deleted {len(objects_to_delete)} files from S3: {process_name}")

    except Exception as e:
        print(f"❌ Error cleaning up S3 folder {process_name}: {e}")
    
def cleanup_local_process_folder(process_name):
            """
            Clean up local process folder after S3 upload.
            """
            local_folder = os.path.join(os.getcwd(), settings.AUDIT_FILES_LOCATION, process_name)
            if os.path.exists(local_folder):
                try:
                    remove_dir_recursive(local_folder)
                    print(f"🧹 Cleaned up local folder: {process_name}")
                except Exception as e:
                    print(f"⚠️ Could not clean local folder {process_name}: {e}")